import openpyxl
import random

wb=openpyxl.Workbook()  # new excel file create
ws=wb.active

ws.title="data"

ws.append(['No.','과목','중간','기말','점수합계'])

name_list=['국어','물리','수학','생물', '영어',' 화학']
           
for i in range(random.randint(5,10)): # 매출 데이터(행) 5~10개
    name=random.choice(name_list)
    if name=="국어":
        score_1 = 80
        score_2 = 80
    elif name=="물리":
        score_1 = 30
        score_2 = 80
    elif name=="수학":
        score_1 = 90
        score_2 = 70
    elif name=="생물":
        score_1 = 80
        score_2 = 40 
    elif name=="영어":
        score_1 = 80
        score_2 = 90 
    elif name=="화학":
        score_1 = 60
        score_2 = 50      
    ws.append([i+1, name, score_1, score_2, random.randint(1,5), f'=C{i+2}*D{i+2}'])
    
wb.save("save_score.xlsx")  # save excel file

spf='save_score.xlsx'

wb=openpyxl.load_workbook(spf, data_only=True)  # load excel file 
ws=wb['data']

for x in range(1, 10):
    for y in range(1, 6):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()